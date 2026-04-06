from openpyxl import Workbook
import random
from datetime import datetime, timedelta

# Crear libro
wb = Workbook()

# Función para fechas aleatorias
def random_date(start, end):
    delta = end - start
    return start + timedelta(days=random.randint(0, delta.days))

start_date = datetime(2025, 1, 1)
end_date = datetime(2026, 4, 1)

# =========================
# HOJA 1: INFRAESTRUCTURA
# =========================
ws1 = wb.active
ws1.title = "Infraestructura"

ws1.append(["ID", "Nombre", "Tipo", "Ubicación", "Estado", "Fecha Registro"])

tipos = ["Puente", "Carretera", "Edificio", "Hospital", "Escuela"]
estados = ["Bueno", "Regular", "Malo"]
ubicaciones = ["Cancún", "Playa del Carmen", "Tulum", "Chetumal"]

for i in range(1, 201):  # 1000 registros
    ws1.append([
        i,
        f"Infraestructura_{i}",
        random.choice(tipos),
        random.choice(ubicaciones),
        random.choice(estados),
        random_date(start_date, end_date).strftime("%Y-%m-%d")
    ])

# =========================
# HOJA 2: PRESUPUESTOS
# =========================
ws2 = wb.create_sheet(title="Presupuestos")

ws2.append(["ID", "Proyecto", "Monto Asignado", "Monto Usado", "Fecha Inicio", "Fecha Fin"])

for i in range(1, 1001):
    asignado = random.randint(100000, 5000000)
    usado = random.randint(50000, asignado)

    fecha_inicio = random_date(start_date, end_date)
    fecha_fin = fecha_inicio + timedelta(days=random.randint(30, 365))

    ws2.append([
        i,
        f"Proyecto_{i}",
        asignado,
        usado,
        fecha_inicio.strftime("%Y-%m-%d"),
        fecha_fin.strftime("%Y-%m-%d")
    ])

# =========================
# HOJA 3: OBRAS
# =========================
ws3 = wb.create_sheet(title="Obras")

ws3.append(["ID", "Nombre Obra", "Tipo", "Responsable", "Estado", "Avance (%)"])

responsables = ["Ing. López", "Ing. Pérez", "Ing. García", "Ing. Ramírez"]
estado_obra = ["Planeado", "En progreso", "Finalizado"]

for i in range(1, 201):
    ws3.append([
        i,
        f"Obra_{i}",
        random.choice(["Construcción", "Mantenimiento", "Rehabilitación"]),
        random.choice(responsables),
        random.choice(estado_obra),
        random.randint(0, 100)
    ])

# =========================
# HOJA 4: INCIDENCIAS
# =========================
ws4 = wb.create_sheet(title="Incidencias")

ws4.append(["ID", "Descripción", "Ubicación", "Fecha", "Gravedad", "Estado"])

gravedad = ["Baja", "Media", "Alta", "Crítica"]
estado_inc = ["Pendiente", "En proceso", "Resuelto"]

for i in range(1, 201):  
    ws4.append([
        i,
        f"Incidencia_{i}",
        random.choice(ubicaciones),
        random_date(start_date, end_date).strftime("%Y-%m-%d"),
        random.choice(gravedad),
        random.choice(estado_inc)
    ])

# Guardar archivo
wb.save("test2.xlsx")

print("✅ Archivo Excel generado correctamente")